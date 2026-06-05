from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_korean_excel_summary(manifest_jsonl: str | Path, qc_report_json: str | Path, out_xlsx: str | Path) -> Path:
    """v4 FDS 산출물을 한국어 Excel 리포트로 정리합니다.

    이 리포트는 사용자가 요구한 “한국어 엑셀 정리”를 자동화하기 위한 산출물입니다.
    데이터 생성 품질, 문서 종류, 사기 유형, visual cluster, QC gate를 한 파일에 모아
    대량 생성 후 어떤 학습 데이터가 만들어졌는지 빠르게 검토할 수 있게 합니다.
    """

    manifest_path = Path(manifest_jsonl)
    qc_path = Path(qc_report_json)
    rows = [json.loads(line) for line in manifest_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    qc = json.loads(qc_path.read_text(encoding="utf-8"))
    out = Path(out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    _write_table(
        ws,
        ["항목", "값"],
        [
            ["총 manifest 행 수", len(rows)],
            ["청구 bundle 수", len({row.get("claim_id") for row in rows})],
            ["문서 유형 수", len({row.get("document_type") for row in rows})],
            ["visual cluster 수", len({row.get("visual_cluster_id") for row in rows if row.get("visual_cluster_id")})],
            ["NO 정상/양성 hard negative 행", sum(1 for row in rows if row.get("label_family") == "NO")],
            ["AF 방어용 counterfactual 행", sum(1 for row in rows if row.get("label_family") == "AF")],
            ["Real Image 파생 행", sum(1 for row in rows if row.get("derived_from_real_profile"))],
            ["QC 통과", str(qc.get("quality_gate", {}).get("pass"))],
            ["tamper mask 생성", "비활성화" if qc.get("quality_gate", {}).get("mask_generation_disabled") else "활성"],
            ["개인정보/실식별자 leak findings", len(qc.get("quality_gate", {}).get("privacy_leakage_findings", []))],
        ],
    )

    _counter_sheet(wb, "문서유형별", rows, "document_type")
    _counter_sheet(wb, "사기유형별", rows, "attack_family")
    _counter_sheet(wb, "사유코드별", _explode_reason_codes(rows), "reason_code")
    _counter_sheet(wb, "클러스터별", rows, "visual_cluster_id")
    _counter_sheet(wb, "라벨패밀리별", rows, "label_family")

    qc_ws = wb.create_sheet("QC")
    gate = qc.get("quality_gate", {})
    _write_table(
        qc_ws,
        ["QC 항목", "값"],
        [
            ["layout_overflow_count", gate.get("layout_overflow_count")],
            ["critical_truncated_fields", json.dumps(gate.get("critical_truncated_fields", []), ensure_ascii=False)],
            ["privacy_leakage_findings", json.dumps(gate.get("privacy_leakage_findings", []), ensure_ascii=False)],
            ["split_leakage_pass", gate.get("split_leakage_pass")],
            ["benign_conditions_not_fraud", gate.get("benign_conditions_not_fraud")],
            ["mask_generation_disabled", gate.get("mask_generation_disabled")],
            ["pass", gate.get("pass")],
        ],
    )

    taxonomy_ws = wb.create_sheet("FDS탐지맥락")
    _write_table(
        taxonomy_ws,
        ["분류", "사유코드", "탐지해야 하는 이유", "대표 조작 필드"],
        [
            ["금액 변조/과청구", "R101/R102/R103/R301", "영수증·세부내역·청구서 금액 불일치와 과다 청구를 탐지", "total_amount, claimed_amount, line_item_amount"],
            ["진단명/질병코드 변조", "R201/R202/R203", "보장 유리한 진단명·질병코드로 바뀐 cross-document inconsistency 탐지", "diagnosis_name, disease_code"],
            ["약품/처방 변조", "R501/R502/R503/R504", "처방전과 약제비영수증의 약품·수량·일수·금액 불일치 탐지", "drug_name, dosage_days, drug_total"],
            ["중복청구/재사용", "R401/R402/R403/R805", "같은 진료 episode 또는 이미지/문서 재사용을 반복 청구로 탐지", "receipt_no, claim_id, perceptual_hash"],
            ["기관/환자 불일치", "R701/R702/R703/R704/R705", "문서 소유자·의료기관·약국·계약 피보험자 불일치 탐지", "patient_token, provider_token"],
            ["문서 이미지 조작", "R801/R802/R803/R804", "핵심 필드 주변의 국소 편집, OCR/레이아웃 불연속, 템플릿 불일치 탐지", "field_bbox, image_forensic_signal"],
        ],
    )

    for sheet in wb.worksheets:
        _autosize(sheet)
    wb.save(out)
    return out


def _explode_reason_codes(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    exploded: list[dict[str, str]] = []
    for row in rows:
        codes = row.get("reason_codes") or []
        if isinstance(codes, str):
            codes = [codes]
        for code in codes:
            exploded.append({"reason_code": code})
    return exploded


def _counter_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], key: str) -> None:
    ws = wb.create_sheet(title)
    counts = Counter(str(row.get(key) or "none") for row in rows)
    data = [[name, count] for name, count in sorted(counts.items())]
    _write_table(ws, [key, "건수"], data)


def _write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value).alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = max_len
